# -*- coding: utf-8 -*-
"""
Зависимости:
    pip install openpyxl xlrd==2.0.1 pyxlsb
"""

import sys
import os
import re
import math
import time
import shutil
import zipfile
import gc
from datetime import datetime, date, timedelta

from openpyxl import load_workbook
from py.TEXT_HANDLERS.lemma import LEGITIMAZE

# Добавляем папку py в sys.path
sys.path.append(os.path.join(os.path.dirname(__file__), 'py'))

DEFAULT_PATH = r"Z:\\"

# ---------- утилиты ----------

def _is_lockfile(name: str) -> bool:
    """Excel/Office создаёт рядом временные файлы с префиксом '~$' — их надо пропускать."""
    return os.path.basename(name).startswith("~$")

def _row_to_text(cells):
    """Собирает строку из ячеек, пропуская None/NaN, даты -> ISO."""
    parts = []
    for c in cells:
        if c is None:
            continue
        if isinstance(c, float) and math.isnan(c):
            continue
        if isinstance(c, (datetime, date)):
            s = c.isoformat()
        else:
            s = str(c)
        s = s.strip()
        if s:
            parts.append(s)
    return " ".join(parts)

def _maybe_excel_serial_to_date(v):
    """Осторожная эвристика для .xlsb: Excel-сериал → ISO-дата (не трогаем обычные числа)."""
    if isinstance(v, (int, float)):
        fv = float(v)
        if 59 < fv < 60000:
            try:
                dt = datetime(1899, 12, 30) + timedelta(days=fv)
                return dt.date().isoformat()
            except Exception:
                return v
    return v

def _size_stable(path, checks=4, interval=0.4):
    """Ждём, пока размер файла 4 раза подряд не изменится (для .xls/.xlsb/SMB)."""
    last = -1
    stable = 0
    while stable < checks:
        try:
            cur = os.path.getsize(path)
        except OSError:
            stable = 0
            time.sleep(interval)
            continue
        if cur == last:
            stable += 1
        else:
            stable = 0
            last = cur
        time.sleep(interval)
    return True

def _xlsx_zip_ready(path) -> bool:
    """Проверяем, что .xlsx/.xlsm — валидный ZIP (файл дописан)."""
    try:
        with zipfile.ZipFile(path) as zf:
            if '[Content_Types].xml' not in zf.namelist():
                return False
            return zf.testzip() is None
    except (zipfile.BadZipFile, PermissionError, OSError):
        return False

# ---------- ридеры ----------

def _read_xlsx_like(file_path: str) -> str:
    text_parts = []
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sh in wb.worksheets:
            for row in sh.iter_rows(values_only=True):
                row_text = _row_to_text(row)
                if row_text:
                    text_parts.append(row_text)
    finally:
        # критично: закрыть, иначе файл «держится» и не перемещается на Windows/SMB
        wb.close()
    return " ".join(text_parts)

def _read_xls(file_path: str) -> str:
    import xlrd  # xlrd==2.0.1
    from xlrd import XL_CELL_DATE
    from xlrd.xldate import xldate_as_datetime

    text_parts = []
    wb = xlrd.open_workbook(file_path, formatting_info=False)
    try:
        dm = wb.datemode
        for sh in wb.sheets():
            for r in range(sh.nrows):
                row_vals = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    val = cell.value
                    if cell.ctype == XL_CELL_DATE:
                        try:
                            val = xldate_as_datetime(val, dm)
                        except Exception:
                            pass
                    row_vals.append(val)
                row_text = _row_to_text(row_vals)
                if row_text:
                    text_parts.append(row_text)
    finally:
        # у xlrd нет close(), но есть release_resources
        try:
            wb.release_resources()
        except Exception:
            pass
        del wb
    return " ".join(text_parts)

def _read_xlsb(file_path: str) -> str:
    from pyxlsb import open_workbook as open_xlsb
    text_parts = []
    with open_xlsb(file_path) as wb:
        for name in wb.sheets:
            with wb.get_sheet(name) as sh:
                for row in sh.rows():
                    row_values = [_maybe_excel_serial_to_date(c.v) for c in row]
                    row_text = _row_to_text(row_values)
                    if row_text:
                        text_parts.append(row_text)
    return " ".join(text_parts)

def read_excel_any(file_path: str) -> str:
    """
    Базовый ридер: .xlsx/.xlsm через openpyxl, .xls через xlrd, .xlsb через pyxlsb.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return _read_xlsx_like(file_path)
    elif ext == '.xls':
        return _read_xls(file_path)
    elif ext == '.xlsb':
        return _read_xlsb(file_path)
    else:
        raise ValueError(f"Неподдерживаемое расширение: {ext}")

def read_excel_any_robust(file_path: str, timeout_s: float = 60.0) -> str:
    """
    Робастная обёртка: ждём, пока Excel-файл действительно готов к чтению.
      - .xlsx/.xlsm: проверяем валидность ZIP (файл дописан)
      - .xls/.xlsb: ждём стабилизацию размера
      - ретраим BadZipFile/PermissionError до timeout_s
    """
    ext = os.path.splitext(file_path)[1].lower()
    deadline = time.time() + timeout_s

    if ext in ('.xls', '.xlsb'):
        _size_stable(file_path, checks=4, interval=0.4)

    while True:
        try:
            if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                if not _xlsx_zip_ready(file_path):
                    raise zipfile.BadZipFile("xlsx not ready")
            return read_excel_any(file_path)
        except (zipfile.BadZipFile, PermissionError, OSError) as e:
            if time.time() >= deadline:
                raise
            time.sleep(0.5)

# ---------- остальная логика ----------

def extract_bucket(file_name: str) -> str:
    """Возвращает фамилию/номер папки по имени файла, игнорируя префикс 'Битый[_ren]_'."""
    base = os.path.basename(file_name)
    name_wo_ext = os.path.splitext(base)[0]
    name_wo_prefix = re.sub(r'^(?:Битый(?:_ren)?_)', '', name_wo_ext, flags=re.IGNORECASE)
    if name_wo_prefix.startswith('(') and ')' in name_wo_prefix:
        return name_wo_prefix[1:name_wo_prefix.index(')')]
    return name_wo_prefix.split('_')[0].strip('()')

def move_to_ready(file_path: str, file_name: str, bucket: str = None):
    """Кладёт файл в Z:\\<bucket>\\Готовые (с ретраями на случай WinError 32)."""
    if bucket is None:
        bucket = extract_bucket(file_name)
    ready_dir = os.path.join(DEFAULT_PATH, bucket, "Готовые")
    os.makedirs(ready_dir, exist_ok=True)
    dst_file = os.path.join(ready_dir, file_name)

    for i in range(6):
        try:
            shutil.move(file_path, dst_file)
            print(f"Файл {file_name} перемещён в {ready_dir}")
            return
        except Exception as e:
            if i == 5:
                print(f"Ошибка при переносе файла: {e}")
            time.sleep(0.5)

def EXCEL_TEXT_EXTRACTOR(file_names, excel_repos):
    """
    На вход: имена файлов и путь к каталогу.
    На выход: { file_name: [lemmas...] }.
    Битые/нечитаемые файлы переименовываются в 'Битый_ren_*' и уезжают в 'Готовые'.
    """
    texts = {}
    for file in file_names:
        # Пропускаем помеченные и офисные lock-файлы
        if "_ren_" in file or _is_lockfile(file):
            continue

        file_path = os.path.join(excel_repos, str(file))
        bucket = extract_bucket(file)

        try:
            text_str = read_excel_any_robust(file_path, timeout_s=60.0)
            lemmas = [w for w in LEGITIMAZE(text_str) if w != "nan"]
            texts[file] = lemmas
            gc.collect()  # на больших пачках снижает шанс «зацепов» дескрипторов
        except Exception as e:
            print(f"Ошибка при извлечении текста ({file}): {e}")
            broken_name = "Битый_ren_" + str(file)
            broken_path = os.path.join(excel_repos, broken_name)
            try:
                os.rename(file_path, broken_path)
            except Exception as e_rename:
                print(f"Ошибка при переименовании битого файла: {e_rename}")
                broken_path = file_path
                broken_name = file
            move_to_ready(broken_path, broken_name, bucket)

    return texts
